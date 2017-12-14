#!/usr/bin/env python
# -*- coding: utf-8 -*-



from openpyxl import load_workbook
import json

def save(htmlContent,json_output):
        '''Coverts a dictionary into a json file.'''
        f = open(json_output, 'w')
        # content = json.dumps(''.join(self.allContent), ensure_ascii=False);
        content=''.join(htmlContent)
        print ("dictionary: " + content)
        f.write(content.encode("utf-8"))
        f.close()

if (__name__ == "__main__"):
    sourceMockPath = "C:\\Users\\g8876\\Desktop\\month\\test.xlsx";
    destMockPath = "C:\\Users\\g8876\\Desktop\\bug.html";
    wb = load_workbook(sourceMockPath)
    print"sheetName: " + json.dumps(wb.sheetnames)
    allContent = [];
    if (wb.sheetnames.__len__() > 0):
        print "first sheetname: " + wb.sheetnames[0]
        sheet = wb.get_sheet_by_name(wb.sheetnames[0])
        print "max_row: "+ str(sheet.max_row)

        for i in range(2, sheet.max_row):
            bugId = "#" + sheet["A" + str(i)].value
            bugHttp = "http://cc.pm.netease.com:8082/issues/" + str(sheet["A" + str(i)].value)
            bugTitle = sheet["G" + str(i)].value;
            print "bugId: " + bugId + " title: " + bugTitle + " bugHttp: " + bugHttp
            allContent.append('<p>')
            allContent.append(bugTitle.strip())
            for i in range(1, 6):
                allContent.append("&nbsp;")
            allContent.append('<a')
            allContent.append(' href="')
            allContent.append(bugHttp)
            allContent.append('"')
            allContent.append(">")
            allContent.append(bugId.strip())
            allContent.append('</a>')
            allContent.append('</p>')
        save(allContent, destMockPath)
